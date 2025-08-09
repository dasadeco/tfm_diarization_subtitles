from enum import Enum
from io import FileIO
import string
import argparse
import os
import logging
import math
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import fills, PatternFill
from pyannote.core import Annotation, Segment
from pyannote.metrics.diarization import DiarizationErrorRate, DiarizationCompleteness, DiarizationCoverage, DiarizationPurity, DiarizationHomogeneity, \
    DiarizationPurityCoverageFMeasure, GreedyDiarizationErrorRate, JaccardErrorRate
from pyannote.metrics.detection import DetectionErrorRate, DetectionAccuracy, DetectionCostFunction, DetectionPrecision, DetectionRecall, DetectionPrecisionRecallFMeasure
from pyannote.metrics.segmentation import SegmentationCoverage, SegmentationPurity, SegmentationPurityCoverageFMeasure, SegmentationPrecision, SegmentationRecall
from pyannote.metrics.identification import IdentificationErrorRate, IdentificationPrecision, IdentificationRecall

RTTM = "rttm"
RTTM_REF = "rttm_ref"
COLLAR = 0.
EXECUTION_TIME_FILE = "exec_time.txt"
EXECUTION_NEMO_TIME_FILE = "NEMO_exec_time.txt"
EXECUTION_PYANNOTE_TIME_FILE = "PYANNOTE_exec_time.txt"  
EXECUTION_SPEECHBRAIN_TIME_FILE = "SPEECHBRAIN_exec_time.txt"  


class DatasetEnum(Enum):
    canaluned = "canal.uned"
    fiapas = "FIAPAS"
    fundaciononce = "Fundacion ONCE"
    one_speaker= "1-speaker"

class MetricsEnum(Enum):
    # First Step
  DetCost = "Detection Cost Function" 
  DetER = "Detection Error Rate"    
  DetAcc = "Detection Accuracy"
  DetPrec= "Detection Precision" 
  DetRec= "Detection Recall"
  DetFMeas= "Detection F-Measure" 
    # Second Step
  SegPur = "Segmentation Purity"
  SegCover = "Segmentation Coverage"
  SegFMeas = "Segmentation F-Measure"
  SegPrec = "Segmentation Precision"
  SegRec = "Segmentation Recall"
    # Third Step
  DER = "Diarization Error Rate"
  DiariCompl = "Diarization Completeness"
  DiariCover = "Diarization Coverage"
  DiariHomog = "Diarization Homogeneity"
  DiariPur = "Diarization Purity"
  DiariFMeas = "Diarization Purity/Coverage F-Measure"
  GreedyDER = "Greedy Diarization Error Rate"
  JER = "Jaccard Error Rate"
    # Fourth Step
  #IER = "Identification Error Rate"
  #IdentPrec = "Identification Precision" 
  #IdentRec = "Identification Recall"
    # Performance
  RTF = "Real Time Factor"  
  
    
class PipelineEnum(Enum):
    PYANNOTE ='Pyannote'
    NEMO ='NeMo' 
    SPEECHBRAIN = 'SpeechBrain'   
    
class PipelineVersions(Enum):
    V2_1 ='speaker-diarization@2.1'
    #V3_0 ='speaker-diarization-3.0'
    V3_1 ='speaker-diarization-3.1'
    _ = 'NA'
    
class VAD_Models(Enum):
    ORACLE = 'oracle_vad'
    MARBLE = 'vad_multilingual_marblenet'
    _ = 'NA'
    
class Speaker_Models(Enum):
    TITANET_LARGE = "titanet_large" 
    ECAPA_TDNN = "ecapa_tdnn" 
    SPEAKER_VERIF = "speakerverification_speakernet"    
    TITANET_SMALL = "titanet_small"    
    _ = 'NA'


class MetricsByAudioFile():    
    def __init__(self, rttm_file:str, segmentation_model:str, vad_model:str, embedding_model:str, metrics_map:dict, dataset=DatasetEnum):
        self.rttm_file = rttm_file
        self.segmentation_model = segmentation_model
        self.vad_model = vad_model
        self.embedding_model = embedding_model
        self.metrics_map = metrics_map        
        if type(dataset) == DatasetEnum:
            self.dataset = dataset.name         
        else:
            self.dataset = dataset

total_metrics:list[MetricsByAudioFile] = []

def _check_in_list(tuplas_list, new_tupla):
    for tupla in tuplas_list:
        if tupla==new_tupla:
            return True
    return False    

def _buscar_by_extension_in_dataset_2_niveles(path, extension):
    resultados = []
    for carpeta_actual, carpetas, _ in os.walk(path):
        for carpeta in carpetas:
           for subcarpeta_actual, _, archivos in os.walk(os.path.join(path, carpeta)): 
            nombre_subcarpeta = os.path.basename(subcarpeta_actual)
            for archivo in archivos:
                if archivo.lower().endswith(extension):
                    new_tupla = (archivo, nombre_subcarpeta, carpeta)
                    if not _check_in_list(resultados, new_tupla):
                        resultados.append(new_tupla)
    return resultados


class MetricsCalculator():
        
    def __init__(self, hypotheses_path, reference_path, metrics_list, collar, skip_overlap):
        self.hypotheses_path = hypotheses_path
        self.reference_path = reference_path
        self.metrics_list = metrics_list
        self.collar = collar
        self.skip_overlap = skip_overlap
        
        logs_path = os.path.join(self.hypotheses_path, os.path.pardir, "logs")
        if not os.path.exists(logs_path):
            os.makedirs( logs_path, exist_ok=True )        
        logging.basicConfig(filename=f'{logs_path}/metrics.log', force=True,
                            encoding='utf-8', level=logging.DEBUG, format='%(asctime)s %(message)s', datefmt='%Y-%m-%d %H:%M:%S')            
        self.logger = logging.getLogger(__name__)
        
        if self.collar is not None and self.collar > 0.5:
            self.logger.warning("Expected collar < 0.500")
                                        
        if self.metrics_list.lower() == 'all':
            self.metrics_list= ''
            for me in MetricsEnum:
                self.metrics_list += me.name + ','
            self.metrics_list = self.metrics_list[:-1]         
        
        
    def calculate_and_write_metrics(self):
        tuplas_hyp = _buscar_by_extension_in_dataset_2_niveles(self.hypotheses_path, ".rttm")                                                                         
        for tupla_hyp in tuplas_hyp:
            rttm_hyp_file = tupla_hyp[0]               
            combined_model_subfolder = tupla_hyp[1]
            if tupla_hyp[1] == tupla_hyp[2]:
                dataset_subfolder_path = os.path.join(self.hypotheses_path, '.')
            else:    
                dataset_subfolder_path = os.path.join(self.hypotheses_path, tupla_hyp[2])
            pipeline = PipelineEnum.PYANNOTE.name if combined_model_subfolder.startswith(PipelineEnum.PYANNOTE.value) \
                else PipelineEnum.NEMO.name if combined_model_subfolder.startswith(PipelineEnum.NEMO.value) \
                else PipelineEnum.SPEECHBRAIN.name
            self.executeMetrics(self.metrics_list.split(','), self.hypotheses_path, dataset_subfolder_path, combined_model_subfolder, \
                rttm_hyp_file, self.reference_path, pipeline, self.collar, self.skip_overlap)

        self.write_metrics()
            
        
  
    def executeMetrics(self, metrics:list, rttms_hyp_path, dataset_subfolder_path, combin_model_subfold, rttm_file, rttms_ref_path, pipeline, collar=COLLAR, skip_overlap=False):
        
        def _create_annot(file:FileIO, annot:Annotation)->Annotation:              
            for line in file:
                if line.strip() != "":
                    data_list = line.split()
                    annot[Segment( math.trunc(float(data_list[3])*100),     
                        math.trunc(float(data_list[3])*100) + math.trunc(float(data_list[4])*100))] = data_list[7]
            return annot
                                                
        dataset = os.path.basename(dataset_subfolder_path)
        hyp_rttm_file_path = os.path.join(dataset_subfolder_path, combin_model_subfold, rttm_file)
        ref_rttm_file_path = os.path.join(rttms_ref_path, dataset, rttm_file)
        hypothesis, reference = None, None
        if os.path.exists(hyp_rttm_file_path):    
            hypothesis = Annotation(rttm_file, combin_model_subfold)
            with open(hyp_rttm_file_path, 'r') as hyp_file:
                hypothesis = _create_annot(hyp_file, hypothesis)
            hyp_file.close()    
        
        if os.path.exists(ref_rttm_file_path):    
            reference = Annotation(rttm_file, RTTM_REF)
            with open(ref_rttm_file_path, 'r') as ref_file:
                reference = _create_annot(ref_file, reference)
            ref_file.close()    
        if collar is None:
            collar = 0.0 
        metrics_map = {}
                                        
        for metric in metrics:
            metric = metric.strip()        
            match metric:
                case MetricsEnum.DetAcc.name : metrics_map[ MetricsEnum.DetAcc.value] = DetectionAccuracy(collar, skip_overlap)(reference, hypothesis) if hypothesis is not None and reference is not None else 'NA'
                case MetricsEnum.DetCost.name : metrics_map[ MetricsEnum.DetCost.value] = DetectionCostFunction(collar, skip_overlap)(reference, hypothesis) if hypothesis is not None and reference is not None else 'NA'
                case MetricsEnum.DetER.name : metrics_map[ MetricsEnum.DetER.value] = DetectionErrorRate(collar, skip_overlap)(reference, hypothesis) if hypothesis is not None and reference is not None else 'NA'
                case MetricsEnum.DetPrec.name : metrics_map[ MetricsEnum.DetPrec.value] = DetectionPrecision(collar, skip_overlap)(reference, hypothesis) if hypothesis is not None and reference is not None else 'NA'
                case MetricsEnum.DetRec.name : metrics_map[ MetricsEnum.DetRec.value] = DetectionRecall(collar, skip_overlap)(reference, hypothesis) if hypothesis is not None and reference is not None else 'NA'
                case MetricsEnum.DetFMeas.name : metrics_map[ MetricsEnum.DetFMeas.value] = DetectionPrecisionRecallFMeasure(collar, skip_overlap)(reference, hypothesis) if hypothesis is not None and reference is not None else 'NA'
                case MetricsEnum.SegPur.name : metrics_map[ MetricsEnum.SegPur.value] = SegmentationPurity(collar)(reference, hypothesis) if hypothesis is not None and reference is not None else 'NA'
                case MetricsEnum.SegCover.name : metrics_map[ MetricsEnum.SegCover.value] = SegmentationCoverage(collar)(reference, hypothesis) if hypothesis is not None and reference is not None else 'NA'
                case MetricsEnum.SegFMeas.name : metrics_map[ MetricsEnum.SegFMeas.value] = SegmentationPurityCoverageFMeasure(collar)(reference, hypothesis) if hypothesis is not None and reference is not None else 'NA'
                case MetricsEnum.DER.name : metrics_map[ MetricsEnum.DER.value] = DiarizationErrorRate(collar, skip_overlap)(reference, hypothesis) if hypothesis is not None and reference is not None else 'NA'
                case MetricsEnum.DiariCompl.name : metrics_map[ MetricsEnum.DiariCompl.value] = DiarizationCompleteness(collar, skip_overlap)(reference, hypothesis) if hypothesis is not None and reference is not None else 'NA'
                case MetricsEnum.DiariPur.name : metrics_map[ MetricsEnum.DiariPur.value] = DiarizationPurity(collar)(reference, hypothesis) if hypothesis is not None and reference is not None else 'NA'            
                case MetricsEnum.DiariCover.name : metrics_map[ MetricsEnum.DiariCover.value] = DiarizationCoverage(collar)(reference, hypothesis) if hypothesis is not None and reference is not None else 'NA'
                case MetricsEnum.DiariFMeas.name : metrics_map[ MetricsEnum.DiariFMeas.value] = DiarizationPurityCoverageFMeasure(collar)(reference, hypothesis) if hypothesis is not None and reference is not None else 'NA'            
                case MetricsEnum.DiariHomog.name : metrics_map[ MetricsEnum.DiariHomog.value] = DiarizationHomogeneity(collar, skip_overlap)(reference, hypothesis) if hypothesis is not None and reference is not None else 'NA'                        
                case MetricsEnum.GreedyDER.name : metrics_map[ MetricsEnum.GreedyDER.value] = GreedyDiarizationErrorRate(collar, skip_overlap)(reference, hypothesis) if hypothesis is not None and reference is not None else 'NA'
                case MetricsEnum.JER.name : metrics_map[ MetricsEnum.JER.value] = JaccardErrorRate(collar, skip_overlap)(reference, hypothesis) if hypothesis is not None and reference is not None else 'NA'
                #case MetricsEnum.IER.name : metrics_map[ MetricsEnum.IER.value] = IdentificationErrorRate(collar)(reference, hypothesis) if hypothesis is not None and reference is not None else 'NA'
                #case MetricsEnum.IdentPrec.name : metrics_map[ MetricsEnum.IdentPrec.value] = IdentificationPrecision(collar)(reference, hypothesis) if hypothesis is not None and reference is not None else 'NA'
                #case MetricsEnum.IdentRec.name : metrics_map[ MetricsEnum.IdentRec.value] = IdentificationRecall(collar)(reference, hypothesis) if hypothesis is not None and reference is not None else 'NA'
                
                #La métrica de rendimiento lleva un proceso totalmente distinto
                case MetricsEnum.RTF.name : metrics_map[MetricsEnum.RTF.value] = self._calcula_ratio(rttms_hyp_path, dataset_subfolder_path, combin_model_subfold, rttm_file, pipeline) if hypothesis is not None else 'NA'
        mbaf = MetricsByAudioFile(rttm_file, combin_model_subfold.split('__')[1].split('+')[0], VAD_Models._.value, combin_model_subfold.split('__')[1].split('+')[1], metrics_map, dataset) if pipeline == PipelineEnum.PYANNOTE.name \
            else MetricsByAudioFile(rttm_file, PipelineVersions._.value, combin_model_subfold.split('__')[1].split('+')[0], combin_model_subfold.split('__')[1].split('+')[1], metrics_map, dataset) if pipeline == PipelineEnum.NEMO.name \
            else None    
        total_metrics.append(mbaf)

    def write_metrics(self):    
        index = [ i for i, _ in enumerate(total_metrics)]
        columns = ["Audios", "Datasets", "Pyannote Segmentation Model", "NeMo VAD Model", "P/N Embeddings Models"]
        rttm_files = [ mbaf.rttm_file for mbaf in total_metrics]    
        datasets = [ mbaf.dataset for mbaf in total_metrics]        
        segmentation_models = [ mbaf.segmentation_model for mbaf in total_metrics]
        vad_models = [ mbaf.vad_model for mbaf in total_metrics]
        embeddings_models = [ mbaf.embedding_model for mbaf in total_metrics]
        data = {"Audios": rttm_files, "Datasets": datasets, "Pyannote Segmentation Model": segmentation_models, "NeMo VAD Model": vad_models, "P/N Embeddings Models": embeddings_models}
        list_metrics = [mbaf.metrics_map for mbaf in total_metrics]
        for metrics in list_metrics:
            for metric in metrics.keys():
                if metric not in columns:
                    columns += [metric]
                if not metric in data:
                    data[metric]=[metrics[metric]]
                else: 
                    data[metric].append(metrics[metric])

        metrics_df = pd.DataFrame(data, index = index, columns=columns)  
        metrics_path = os.path.join(self.hypotheses_path, os.path.pardir, "metrics")
        if not os.path.exists(metrics_path):
            os.makedirs( metrics_path, exist_ok=True )          
        export_path = os.path.join(metrics_path, "metrics.xlsx")
        metrics_df.to_excel(export_path,  index=False)
        wb = load_workbook(export_path)
        ws = wb["Sheet1"]   
        ws.column_dimensions['A'].width = 55
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 30
        for letter in string.ascii_uppercase[5:]:
            ws.column_dimensions[letter].width = 27
            
        odd_fill = PatternFill(fills.FILL_PATTERN_LIGHTUP)               
        for number in range(2, len(index)+2):        
            if ws.cell(row=number, column=3).value == 'NA':
                for y in range(1, ws.max_column+1):
                    ws.cell(row=number, column=y).fill = odd_fill
            
        wb.save(export_path)
        
        
    ## Buscamos el archivo en el que está guardado el tiempo de ejecución del archivo de ese dataset usando ese modelo    
    def _calcula_ratio(self, base_rttms_hyp_path, dataset_subfolder_path, combined_model, rttm_file, pipeline:str):        
        rtf = 'NA'   
        exec_file_path = os.path.join(base_rttms_hyp_path, pipeline + '_' + EXECUTION_TIME_FILE)
        dataset = os.path.basename(dataset_subfolder_path)
        with open(exec_file_path, 'r', encoding="utf-8") as exec_file:
            for line in exec_file:
                word = line.rstrip().split(" ")
                if rttm_file == word[0] and combined_model == word[1] and dataset == word[2]:
                    exec_time = word[3]
                    duration = word[4]        
                    break
        if rttm_file == word[0] and combined_model == word[1] and dataset == word[2]:
            rtf = float(exec_time)/float(duration)  # Real-Time Factor
            self.logger.info(f"Ratio de procesamiento del audio {rttm_file.replace('.rttm', '')}: {str(rtf)}")
            print(f"Ratio de procesamiento del audio {rttm_file.replace('.rttm', '')}: {str(rtf)}")          
        return rtf
            

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Pyannote Metrics')
    parser.add_argument('-hp', '--hypotheses_path', type=str, default='E:\Desarrollo\TFM\data\media\rttm', help='Carpeta con los archivos RTTM de hipotesis.')
    parser.add_argument('-rp', '--reference_path', type=str, default='E:\Desarrollo\TFM\subtitles\data\rttm_ref', help='Carpeta con los archivos RTTM de referencia.')
    parser.add_argument('-me', '--metrics_list', type=str, help='Lista de Metricas a aplicar')
    parser.add_argument('-co', '--collar', type=float, help='Collar (Umbral de holgura al principio  al final de cada segmento)')
    parser.add_argument('-so', '--skip_overlap', type=bool, default=False, help='Si se ignora el habla solapada o no')
    args = parser.parse_args()
            
    if os.path.exists(args.hypotheses_path) and os.path.exists(args.reference_path):
        metrics_calc = MetricsCalculator(hypotheses_path=args.hypotheses_path, reference_path=args.reference_path, metrics_list=args.metrics_list, 
                        collar=args.collar, skip_overlap=args.skip_overlap)
        metrics_calc.calculate_and_write_metrics()
    else:
        print("No existe la carpeta de archivos RTTM de hipótesis o la carpeta de archivos de referencia. NO se pueden calcular las métricas.")
            